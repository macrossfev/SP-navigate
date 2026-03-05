#!/usr/bin/env python3
"""Generate an overview map showing all 80 points across 16 days with colored regions."""

import openpyxl
import folium
from folium import DivIcon
import math
from collections import defaultdict

# --- Config ---
PLAN_XLSX = "/root/projects/navigate/route_system/output/最终方案_质心聚类/采样计划总表_最终方案.xlsx"
ADDR_XLSX = "/root/projects/navigate/最终地址列表.xlsx"
OUT_DIR = "/root/projects/navigate/route_system/output/最终方案_质心聚类"
OUT_HTML = f"{OUT_DIR}/全局总览图.html"
OUT_PNG = f"{OUT_DIR}/全局总览图.png"

OUTLIER_NAMES = ["重庆市长寿区新市华府", "重庆市长寿区长风厂"]

COLORS_16 = [
    "#e6194b", "#3cb44b", "#4363d8", "#f58231",
    "#911eb4", "#42d4f4", "#f032e6", "#bfef45",
    "#fabed4", "#469990", "#dcbeff", "#9A6324",
    "#800000", "#aaffc3", "#808000", "#000075",
]

TILE_URL = "https://webrd02.is.autonavi.com/appmaptile?lang=zh_cn&size=1&scale=2&style=8&x={x}&y={y}&z={z}"

CSS_INJECT = '<style>html,body{margin:0;padding:0;width:100%;height:100%;overflow:hidden}.folium-map{position:absolute;top:0;left:0;right:0;bottom:0}</style>'


def calc_zoom(coords, w=1400, h=900):
    lats = [c[0] for c in coords]
    lngs = [c[1] for c in coords]
    lat_range = (max(lats) - min(lats)) * 1.3 or 0.005
    lng_range = (max(lngs) - min(lngs)) * 1.3 or 0.005
    for z in range(18, 5, -1):
        deg_per_px = 360 / (256 * 2 ** z)
        if lng_range / deg_per_px < w and lat_range / deg_per_px < h:
            return z
    return 6


def convex_hull(points):
    """Simple Graham scan for convex hull. points = list of (lat, lng)."""
    pts = sorted(set(points))
    if len(pts) <= 1:
        return pts
    if len(pts) == 2:
        return pts

    def cross(o, a, b):
        return (a[0] - o[0]) * (b[1] - o[1]) - (a[1] - o[1]) * (b[0] - o[0])

    lower = []
    for p in pts:
        while len(lower) >= 2 and cross(lower[-2], lower[-1], p) <= 0:
            lower.pop()
        lower.append(p)
    upper = []
    for p in reversed(pts):
        while len(upper) >= 2 and cross(upper[-2], upper[-1], p) <= 0:
            upper.pop()
        upper.append(p)
    return lower[:-1] + upper[:-1]


def make_circle_points(center, radius_deg, n=36):
    """Generate circle polygon points around center with given radius in degrees."""
    pts = []
    for i in range(n):
        angle = 2 * math.pi * i / n
        lat = center[0] + radius_deg * math.sin(angle)
        lng = center[1] + radius_deg * math.cos(angle)
        pts.append((lat, lng))
    return pts


def main():
    # 1. Read address coordinates
    addr_wb = openpyxl.load_workbook(ADDR_XLSX)
    addr_ws = addr_wb.active
    addr_coords = {}  # address -> (lat, lng)
    for row in addr_ws.iter_rows(min_row=2, values_only=True):
        addr, _, coord_str, _ = row[0], row[1], row[2], row[3]
        if coord_str:
            lng, lat = coord_str.split(",")
            addr_coords[addr] = (float(lat), float(lng))

    # 2. Read plan - group by day
    plan_wb = openpyxl.load_workbook(PLAN_XLSX)
    plan_ws = plan_wb.active
    day_points = defaultdict(list)  # day_num -> [(lat, lng), ...]
    for row in plan_ws.iter_rows(min_row=2, values_only=True):
        day_str = row[0]  # e.g. "第1天"
        coord_str = row[4]  # 坐标
        if day_str and coord_str:
            day_num = int(day_str.replace("第", "").replace("天", ""))
            lng, lat = coord_str.split(",")
            day_points[day_num].append((float(lat), float(lng)))

    # 3. Outlier points
    outlier_coords = {}
    for name in OUTLIER_NAMES:
        if name in addr_coords:
            outlier_coords[name] = addr_coords[name]
        else:
            # fuzzy match
            for k, v in addr_coords.items():
                if name.replace("重庆市长寿区", "") in k:
                    outlier_coords[name] = v
                    break

    # 4. Collect all coords for zoom
    all_coords = []
    for pts in day_points.values():
        all_coords.extend(pts)
    for c in outlier_coords.values():
        all_coords.append(c)

    center_lat = sum(c[0] for c in all_coords) / len(all_coords)
    center_lng = sum(c[1] for c in all_coords) / len(all_coords)
    zoom = calc_zoom(all_coords)

    # 5. Create map
    m = folium.Map(
        location=[center_lat, center_lng],
        zoom_start=zoom,
        tiles=TILE_URL,
        attr="AMap",
    )

    # 6. Draw each day's region and points
    for day_num in sorted(day_points.keys()):
        pts = day_points[day_num]
        color = COLORS_16[(day_num - 1) % 16]

        # Draw points
        for lat, lng in pts:
            folium.CircleMarker(
                location=[lat, lng],
                radius=5,
                color=color,
                fill=True,
                fill_color=color,
                fill_opacity=0.8,
                weight=1,
            ).add_to(m)

        # Draw region polygon
        if len(pts) >= 3:
            hull = convex_hull(pts)
            hull_closed = hull + [hull[0]]
            folium.Polygon(
                locations=hull_closed,
                color=color,
                fill=True,
                fill_color=color,
                fill_opacity=0.15,
                weight=2,
            ).add_to(m)
        elif len(pts) == 2:
            # Draw an ellipse-like shape
            cx = (pts[0][0] + pts[1][0]) / 2
            cy = (pts[0][1] + pts[1][1]) / 2
            r = math.sqrt((pts[0][0] - pts[1][0]) ** 2 + (pts[0][1] - pts[1][1]) ** 2) / 2 + 0.002
            circle_pts = make_circle_points((cx, cy), r)
            folium.Polygon(
                locations=circle_pts,
                color=color,
                fill=True,
                fill_color=color,
                fill_opacity=0.15,
                weight=2,
            ).add_to(m)
        elif len(pts) == 1:
            folium.Circle(
                location=pts[0],
                radius=300,
                color=color,
                fill=True,
                fill_color=color,
                fill_opacity=0.15,
                weight=2,
            ).add_to(m)

        # Label at centroid
        clat = sum(p[0] for p in pts) / len(pts)
        clng = sum(p[1] for p in pts) / len(pts)
        folium.Marker(
            location=[clat, clng],
            icon=DivIcon(
                html=f'<div style="font-size:11px;font-weight:bold;color:{color};text-shadow:-1px -1px 0 #fff,1px -1px 0 #fff,-1px 1px 0 #fff,1px 1px 0 #fff;white-space:nowrap;">第{day_num}天</div>',
                icon_size=(60, 20),
                icon_anchor=(30, 10),
            ),
        ).add_to(m)

    # 7. Outlier markers
    for name, (lat, lng) in outlier_coords.items():
        short = name.replace("重庆市长寿区", "")
        folium.Marker(
            location=[lat, lng],
            icon=DivIcon(
                html=f'<div style="font-size:18px;color:#666;text-align:center;line-height:20px;">✕</div>',
                icon_size=(20, 20),
                icon_anchor=(10, 10),
            ),
        ).add_to(m)
        folium.Marker(
            location=[lat, lng],
            icon=DivIcon(
                html=f'<div style="font-size:10px;color:#666;text-shadow:-1px -1px 0 #fff,1px -1px 0 #fff,-1px 1px 0 #fff,1px 1px 0 #fff;white-space:nowrap;">{short}</div>',
                icon_size=(80, 16),
                icon_anchor=(-5, -5),
            ),
        ).add_to(m)

    # 8. Save HTML with CSS injection
    m.save(OUT_HTML)
    with open(OUT_HTML, "r") as f:
        html = f.read()
    html = html.replace("</head>", CSS_INJECT + "</head>")
    with open(OUT_HTML, "w") as f:
        f.write(html)

    print(f"HTML saved to {OUT_HTML}")

    # 9. Screenshot
    import subprocess
    cmd = [
        "chromium", "--headless", "--no-sandbox", "--disable-gpu",
        "--disable-software-rasterizer",
        f"--screenshot={OUT_PNG}",
        "--window-size=1400,900",
        "--virtual-time-budget=5000",
        f"file://{OUT_HTML}",
    ]
    subprocess.run(cmd, check=True)
    print(f"PNG saved to {OUT_PNG}")


if __name__ == "__main__":
    main()
